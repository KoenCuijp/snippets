# Can you write a Python script for me that does the following?
#
# It loads an excel sheet named "inkomsten.xslx" and does the following:
# - Group all rows on the column "Contactpersoon"
# - Count how many rows this person has, name this "Aantal facturen"
# - Sum the column "Totaal incl. btw" for those rows and name this "Totaal incl. btw"
# - Sum the column "Totaal excl. btw" for those rows and name this "Totaal excl. btw"
# - Output for every person: Aantal facturen and the totals
# - Output the totals summarized over all persons
#
# Then load an excel sheet named "kosten.xslx" and do the following:
# - Take all rows with Categorie=4601 Kantoor kosten maandelijks or (categorie=4800 Softwarekosten and knab in omschrijving.lower())
# - Sum the column Totaal incl. btw for all rows, output as total_monthly_costs_incl_vat
# - Sum the column Totaal excl. btw for all rows, output as total_monthly_costs_excl_vat
# - Take all rows with Categorie 4602 Kantoor kosten eenmalig
# - Sum the column Totaal incl. btw for all rows, output as total_onetime_costs_incl_vat
# - Sum the column Totaal excl. btw for all rows, output as total_onetime_costs_excl_vat
#
# Then output:
# - Total income excl tax - total_monthly_costs_excl_vat (name this profit_on_recurring_invoices_excl_vat)
# - Total income incl tax - total_monthly_costs_incl_vat (name this profit_on_recurring_invoices_incl_vat)
#
# The output:
# - profit_on_recurring_invoices_excl_vat - total_onetime_costs_excl_vat (name this total_profit_excl_vat)
# - profit_on_recurring_invoices_incl_vat - total_onetime_costs_incl_vat (name this total_profit_incl_vat)
#

from openpyxl import load_workbook
from collections import defaultdict

def load_excel_data(filename, sheetname=None):
    """Load data from an Excel sheet into a list of dicts."""
    wb = load_workbook(filename, data_only=True)
    ws = wb[sheetname] if sheetname else wb.active
    rows = list(ws.iter_rows(values_only=True))
    headers = [str(h).strip() for h in rows[0]]
    data = []
    for row in rows[1:]:
        item = {headers[i]: row[i] for i in range(len(headers))}
        data.append(item)
    return data


def process_income(filename):
    data = load_excel_data(filename)
    grouped = defaultdict(lambda: {
        "Aantal facturen": 0,
        "Totaal incl. btw": 0.0,
        "Totaal excl. btw": 0.0
    })

    months = set()

    # Get amount of invoices paid per person
    for row in data:
        month = row["Datum"].month
        if month:
            months.add(month)
        person = str(row.get("Contactpersoon", "")).strip()
        incl = float(row.get("Totaal incl. btw", 0) or 0)
        excl = float(row.get("Totaal excl. btw", 0) or 0)

        grouped[person]["Aantal facturen"] += 1
        grouped[person]["Totaal incl. btw"] += incl
        grouped[person]["Totaal excl. btw"] += excl

    # Add invoices for Koen
    nr_of_months = len(months)
    grouped["Koen Cuijp"]["Aantal facturen"] = nr_of_months
    grouped["Koen Cuijp"]["Totaal incl. btw"] += nr_of_months * 327.91
    grouped["Koen Cuijp"]["Totaal excl. btw"] += nr_of_months * 271.00

    # Print per person
    print("\n=== Inkomsten per huurder ===")
    total_incl = total_excl = total_count = 0
    for person, vals in sorted(grouped.items()):
        print(f"{person.strip()[0:25].ljust(25)} \t {vals['Aantal facturen']} facturen \t {vals['Totaal excl. btw']:.2f} excl. btw")
        total_count += vals['Aantal facturen']
        total_incl += vals['Totaal incl. btw']
        total_excl += vals['Totaal excl. btw']

    print("-----------------------------------------------------------------------------------------------")
    print(f"Totaal inkomsten: {total_excl:.2f} excl. btw")

    return total_incl, total_excl, nr_of_months


def process_costs(filename, nr_of_months):
    data = load_excel_data(filename)

    total_monthly_costs_incl_vat = 0.0
    total_monthly_costs_excl_vat = 0.0
    total_onetime_costs_incl_vat = 0.0
    total_onetime_costs_excl_vat = 0.0

    costs_monthly = {}
    costs_onetime = {}

    for row in data:
        cat = str(row.get("Categorie", "")).strip()
        desc = str(row.get("Omschrijving", "")).lower()
        incl = float(row.get("Totaal incl. btw", 0) or 0)
        excl = float(row.get("Totaal excl. btw", 0) or 0)
        relatie = str(row.get("Relatie", "")).strip()
        date = row.get("Datum")

        # Monthly costs
        if cat == "4601 Kantoor kosten maandelijks" or (cat == "4800 Softwarekosten" and "knab" in desc):
            total_monthly_costs_incl_vat += incl
            total_monthly_costs_excl_vat += excl
            costs_monthly[f"{relatie} - {desc[0:20]} - {date.strftime("%d-%m-%Y")}"] = excl

        # One-time costs
        elif cat == "4602 Kantoor kosten eenmalig":
            total_onetime_costs_incl_vat += incl
            total_onetime_costs_excl_vat += excl
            costs_onetime[desc] = excl

    # Add cleaning costs
    cleaning_costs = nr_of_months * 75.83
    total_monthly_costs_incl_vat += cleaning_costs
    total_monthly_costs_excl_vat += cleaning_costs
    costs_monthly["Schoonmaker"] = cleaning_costs

    print("\n=== Kosten maandelijks ===")
    for desc, amount in sorted(costs_monthly.items()):
        print(f"{desc[:60].ljust(60)} \t {amount:.2f} excl. btw")
    print("-----------------------------------------------------------------------------------------------")
    print(f"Totaal maandelijks excl. btw: {total_monthly_costs_excl_vat:.2f}")

    print("\n=== Kosten eenmalig ===")
    for desc, amount in sorted(costs_onetime.items()):
        print(f"{desc[:60].ljust(60)} \t {amount:.2f} excl. btw")
    print("-----------------------------------------------------------------------------------------------")
    print(f"Totaal eenmalig excl. btw: {total_onetime_costs_excl_vat:.2f}")

    return (
        total_monthly_costs_incl_vat, total_monthly_costs_excl_vat,
        total_onetime_costs_incl_vat, total_onetime_costs_excl_vat
    )


def main():
    total_income_incl, total_income_excl, nr_of_months = process_income("/Users/koencuijp/inkomsten.xlsx")
    (
        monthly_incl, monthly_excl,
        onetime_incl, onetime_excl
    ) = process_costs("/Users/koencuijp/kosten.xlsx", nr_of_months)

    profit_on_recurring_invoices_excl_vat = total_income_excl - monthly_excl
    profit_on_recurring_invoices_incl_vat = total_income_incl - monthly_incl

    total_profit_excl_vat = profit_on_recurring_invoices_excl_vat - onetime_excl
    total_profit_incl_vat = profit_on_recurring_invoices_incl_vat - onetime_incl

    print("\n=== Resultaten ===")
    print(f"Inkomsten maandelijks - Kosten maandelijks: {profit_on_recurring_invoices_excl_vat:.2f}")
    print(f"Inkomsten maandelijks - Kosten maandelijks - Kosten eenmalig: {total_profit_excl_vat:.2f}")


if __name__ == "__main__":
    main()
